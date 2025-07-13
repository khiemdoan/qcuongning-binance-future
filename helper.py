key="7NvUEUX4tnzOja5KQ99gmUG37DQOV9oelvz1akWAr2Zts9X57djRMwbvfgjQoykp"
secret="X9CWCXNsdypjEU8Q0AQaoaqPrcnaX4wpDe5KVxsAfThkVJJAvufiGJ3tb95QqnQC"

key_test = "c21f1bb909318f36de0f915077deadac8322ad7df00c93606970441674c1b39b"
secret_test = "be2d7e74239b2e959b3446b880451cbb4dee2e6be9d391c4c55ae7ca0976f403"
# key api test2
import sys
import time, math

import numpy as np
# import binance_ft
# from binance_ft.um_futures import UMFutures
import openpyxl
import requests
from datetime import datetime, timedelta
import pandas as pd

def over_time(time_string, time_format="%Y-%m-%d %H:%M:%S"):
    given_time = datetime.strptime(time_string, time_format)
    now = datetime.now()
    time_difference = now - given_time
    return time_difference > timedelta(minutes=150)
def post_tele(msg):
    try:
        BOT_TOKEN = '6665726347:AAGJ60eAiJWsurJVf1YrIALjzAWJbrvrZzk'
        CHAT_ID = '942157886'


        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"

        data = {
            "chat_id": CHAT_ID,
            "text": msg,
            "parse_mode":"Markdown"
        }

        response = requests.post(url, data=data)
    except requests.exceptions.SSLError as e:
        print("post_tele", e)

def update_tele(last_command_time, enable_new_order, enable_pnl, indexz, dict_price, askPrice):
    BOT_TOKEN = '6665726347:AAGJ60eAiJWsurJVf1YrIALjzAWJbrvrZzk'

    url = f"https://api.telegram.org/bot{BOT_TOKEN}/getUpdates"
    response = requests.get(url)
    data = response.json()
    
    if "result" in data and len(data["result"]) > 0:
        if "message" not in data["result"][-1].keys():
            return last_command_time, enable_new_order, enable_pnl, indexz, dict_price
        cmd = data["result"][-1]["message"]["text"]
        time = data['result'][-1]['message']['date']
        if time > last_command_time:
            last_command_time = time
            print(f"Chat:{datetime.fromtimestamp(time)}: {cmd}")
            if cmd == "hi":
                post_tele("hello!!!")
            elif cmd == "price":
                post_tele(str(askPrice))
            elif cmd == "off":
                enable_new_order = False
                post_tele("turn off")
            elif cmd == "on":
                enable_new_order = True
                post_tele("turn on")
            elif cmd == "pnl":
                enable_pnl = True
                indexz = 498
                post_tele("trigger pnl: "+str(enable_pnl))
            elif cmd.startswith("rm"):
                idrm = cmd.split(" ")
                if len(idrm):
                    id_rm = int(idrm[1])
                    post_tele(f"remove {id_rm}")
                    if id_rm in dict_price:
                        dict_price[id_rm]['break'] = True
                    else:
                        post_tele(f"not found {id_rm}")
    return last_command_time, enable_new_order, enable_pnl, indexz, dict_price

def is_pin_bar(candle):
    open_price = candle[0]
    high_price = candle[1]
    low_price = candle[2]

    close_price = candle[3]
    body_size = abs(close_price - open_price)
    upper_wick_size = high_price - max(open_price, close_price)
    candle_range = high_price - low_price
    
    # Define Pin Bar conditions
    long_tail = (upper_wick_size > 8 * body_size)
    small_body = body_size <= 0.2 * candle_range
    
    return long_tail and small_body

def is_decrease(candle):
    open_price = candle[0]
    high_price = candle[1]
    low_price = candle[2]

    close_price = candle[3]
    return close_price < open_price

def xlsx_to_nested_dict(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        key = row[0]
        values = row[1:]
        data[key] = {}
        for col_idx, col_value in enumerate(values, start=1):
            data[key][sheet.cell(row=1, column=col_idx).value] = col_value

    return data

def get_commision(ft_order_id, um_futures_client, pair):
    bnb_price = float(um_futures_client.ticker_price("BNBUSDT")['price'])

    
    while True:
        res_query = um_futures_client.query_order(symbol=pair, orderId=ft_order_id, recvWindow=6000)
        if res_query['status'] == "FILLED":
            break
        time.sleep(0.1)

    res_ft_open = um_futures_client.get_account_trades(symbol=pair, orderId=ft_order_id)
    commision_open_future = 0
    for res in res_ft_open:
        if res['commissionAsset']=="USDT":
            commision_open_future += float(res['commission'])
        elif res['commissionAsset']=="BNB":
            commision_open_future += float(res['commission'])*bnb_price
    all_spend = sum([float(res['quoteQty']) for res in res_ft_open])
    all_coin = sum([float(res['qty']) for res in res_ft_open])
    all_pnl = sum([float(res['realizedPnl']) for res in res_ft_open])

    mean_price = all_spend/all_coin
    # if res_ft_open[0]['maker']:
    #     print("--> this order is maker")
    # else:
    #     print("--> this order is taker")
    return commision_open_future, all_spend, all_coin, mean_price, all_pnl

def get_precision(pair, um_futures_client):
    future_info = um_futures_client.exchange_info()['symbols']
    kk = [i['filters'] for i in future_info if i['symbol'] == pair][0]
    precision_qlt = int(-math.log10(float(kk[1]['minQty'])))
    precision_price = int(-math.log10(float(kk[0]['minPrice'])))

    return precision_qlt, precision_price

def get_status_pos(pair, um_futures_client):
    list_risks = um_futures_client.get_position_risk(recvWindow=6000)
    for pos in list_risks:
        if pos['symbol'] == pair:
            mean_price = float(pos['entryPrice'])
            pnl = float(pos["unRealizedProfit"])
            coin_num = abs(float(pos["positionAmt"]))
            return mean_price, pnl, coin_num
    else:
        return 0, 0, 0

def compute_rsi_pd(prices, period=14):
    delta = prices.diff()
    up = delta.clip(lower=0)
    down = -1 * delta.clip(upper=0)
    ema_up = up.ewm(com=period - 1, adjust=False).mean()
    ema_down = down.ewm(com=period - 1, adjust=False).mean()
    rs = ema_up / ema_down
    rsi = 100 - (100 / (1 + rs))
    return rsi

def calculate_rsi_with_ema(x_array, window: int = 14):
    data = x_array[:, 3]
    data = pd.Series(data)

    delta = data.diff()

    # Bước 2: Tách mức tăng (gain) và mức giảm (loss)
    gain = delta.where(delta > 0, 0)
    loss = -delta.where(delta < 0, 0)

    # Bước 3: Tính trung bình động hàm mũ (EMA) cho gain và loss
    avg_gain = gain.ewm(span=window, min_periods=window).mean()
    avg_loss = loss.ewm(span=window, min_periods=window).mean()

    # Bước 4: Tính RS và RSI
    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))

    return np.array(rsi)
# Tính RSI với trung bình động hàm mũ (EMA)
# rsi_ema = calculate_rsi_with_ema(data, window=14)
# print(rsi_ema)
def calculate_rsi(x_array, window: int = 14) -> pd.Series:
    data = x_array[:, 3]
    data = pd.Series(data)
    delta = data.diff()
    
    # Bước 2: Tách riêng mức tăng (gain) và mức giảm (loss)
    gain = delta.where(delta > 0, 0)  # Nếu thay đổi dương thì giữ nguyên, ngược lại bằng 0
    loss = -delta.where(delta < 0, 0)  # Nếu thay đổi âm thì lấy giá trị tuyệt đối, ngược lại bằng 0

    # Bước 3: Tính trung bình động của mức tăng và mức giảm trong `window` ngày
    avg_gain = gain.rolling(window=window, min_periods=1).mean()
    avg_loss = loss.rolling(window=window, min_periods=1).mean()

    # Bước 4: Tính Relative Strength (RS)
    rs = avg_gain / avg_loss

    # Bước 5: Tính RSI
    rsi = 100 - (100 / (1 + rs))

    return np.array(rsi)


def generate_date_list(start_date_str, end_date_str, interval_days=4):
    # Convert string inputs to dat|etime objects
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    
    # Generate list of dates
    date_list = []
    current_date = start_date
    while current_date <= end_date:
        date_list.append(current_date.strftime('%Y-%m-%d'))
        current_date += timedelta(days=interval_days)
    
    return date_list
def get_binance_ohlc_time(symbol, interval, start_str, end_str):
    print(start_str,"->",end_str)
    url = f'https://fapi.binance.com/fapi/v1/klines'
    # url = f'https://api.binance.com/api/v3/klines'

    # url = "https://api.mexc.com/api/v3/klines"
    print(url)


    
    # Convert start and end times to milliseconds since epoch
    start_ts = int(datetime.strptime(start_str, '%Y-%m-%d %H:%M:%S').timestamp() * 1000)
    end_ts = int(datetime.strptime(end_str, '%Y-%m-%d %H:%M:%S').timestamp() * 1000)
    
    params = {
        'symbol': symbol,
        'interval': interval,
        'startTime': start_ts,
        'endTime': end_ts
    }
    response = requests.get(url, params=params)
    if response.status_code!= 200:
        print(f"Error fetching data: {response.status_code} {response.reason}")
        return []
    data = response.json()
    ohlc = []
    
    for item in data:
        ohlc.append({
            'timestamp': datetime.fromtimestamp(item[0] / 1000),
            'open': float(item[1]),
            'high': float(item[2]),
            'low': float(item[3]),
            'close': float(item[4]),
            'volumn': float(item[5])
        })
    
    df = pd.DataFrame(ohlc)
    df.reset_index(drop=True, inplace=True)

    # df.set_index('timestamp', inplace=True)
    return df

def generate_df_klines(start, end, symb, interval, split_day=4):
    list_day = generate_date_list(start, end, split_day)
    list_day.append(end)
    df_all = []
    post_fix= "00"
    # if interval == "15m":
    #     post_fix = "00"
    # elif interval == "4h":
    #     post_fix = "03"
    for i in range(len(list_day)-1):
        df = get_binance_ohlc_time(symb, interval, list_day[i]+f" {post_fix}:05:00", list_day[i+1]+f" {post_fix}:03:00")
        df_all.append(df)
    df_all = pd.concat(df_all,ignore_index=True, axis=0)
    csv_file = f"data_his/{symb}_{start}_{end}_{interval}_fapiv1.csv"
    df_all.to_csv(csv_file, index=False)
    df_all = pd.read_csv(csv_file)
    return df_all, csv_file



